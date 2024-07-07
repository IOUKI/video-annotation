from flask import Blueprint, request, jsonify
from flask.views import MethodView
import os
import openpyxl

router = Blueprint('chart', __name__)

def allowedFile(filename):
    ALLOWED_EXTENSIONS = {'xlsx'}
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

class ChartAPI(MethodView):
    
    def __init__(self) -> None:
        super().__init__()

    def post(self):
        # 檢查是否有上傳檔案
        file = request.files.get('file', None)
        if file is None:
            return jsonify({'message': '未上傳檔案'}), 400
        
        # 檢查檔案格式
        if not allowedFile(file.filename):
            return jsonify({'message': '檔案格式錯誤'}), 400
        
        filePath = os.path.join(os.getcwd(), 'data', 'data.xlsx')

        # 儲存檔案
        file.save(filePath)

        # 讀取檔案
        # header = {
        #     1: 'ID',
        #     2: 'filename',
        #     3: 'start',
        #     4: 'end',
        #     5: 'duration',
        #     6: 'token'
        # }
        time = ''
        process = ''
        yLabel = ''
        dataFrame = openpyxl.load_workbook(filePath)
        dataFrame1 = dataFrame.active
        for index, row in enumerate(dataFrame1.iter_rows()):
            if index == 0:
                continue
            elif index == 1: # first row
                yLabel = row[1].value
                time = f"{row[2].value},{row[3].value}"
                process = row[5].value
            elif index == dataFrame1.max_row - 1: # last row
                time += f",{row[3].value}"
                process += f",{row[5].value},{row[5].value}"
            else:
                time += f",{row[3].value}"
                process += f",{row[5].value}"
        
        data = {
            'time': time,
            'process': process,
            'yLabel': yLabel,
        }

        return jsonify(data), 200

router.add_url_rule('/', view_func=ChartAPI.as_view('ChartAPI'))